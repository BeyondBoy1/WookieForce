from openpyxl import load_workbook
import openpyxl
import wookieforce

#Load an existing workbook to manipulate
workbook_object = openpyxl.load_workbook('newestCWLfileLATEST.xlsx')

#Create a new worksheet object and call in 'Filtered data
worksheet_object1 = workbook_object.create_sheet('fullrosterCWLrole')

'''
All clans, All names, TH, Current location.
'''
async def get_full_roster(client):
	roster = {}
	clans = wookieforce.all_clans()
	for clan_name in clans:
		clan = await client.get_clan(clans[clan_name])
		print('Clan Name:',clan)
		members = await get_members(client,clan)
		roster.update([(clan_name, members)])
	return roster

'''
Testing workbook
'''
#load workbook
ALPHAcwlBUILD = load_workbook(filename = 'newestCWLfileLATEST.xlsx')


#get sheet
#roster = 'ALPHAcwlBUILD.xlsm'.active

#create sheet
clanroster = ALPHAcwlBUILD.create_sheet("clanroster")
clanroster.title = "clanroster"
#so ^ now we have a sheet called clanroster... need to write to it

#get clan data

'''
Get all Clan War League Leagues 
'''
# async def get_clans(client,clan):
# 	clans = []
# 	clan = await client.get_clan('tag')
# 	#get the clan league based on the tag
# 	#append to clans[]
# 	#return clan

'''
Get all Clan Members for a given clan
'''
async def get_members(client,clan):
	members = []
	async for member in clan.get_detailed_members():
		members.append(member)
	return members


'''
Building a table to display member data
'''
async def build_roster_table(client,roster):
	roster_data = 'PlayerTag,PlayerName,Clan Role,TownHall,PlayerLeague,Clan\n'
	for clan in roster:
		players = roster[clan]
		for i in range(len(players)):
			player = players[i]
			#TEST
			roster_data += str(player.tag) + ','
			roster_data += str(player.name) + ','
			roster_data += str(player.role) + ','
			roster_data += str(player.town_hall) + ','
			roster_data += str(player.league) + ','
			roster_data += str(player.clan) + '\n'
			#TEST

	return roster_data


#Save the manipulated workbook
workbook_object.save('newestCWLfileLATEST.xlsx')


